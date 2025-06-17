from openpyxl import load_workbook

wb = load_workbook('sagatave_eksamenam.xlsx',data_only=True)
ws = wb['Lapa_0']
max_row = ws.max_row

total_sum = 0

for row in range(2, max_row + 1):
    client = ws['F' + str(row)].value
    quantity = ws['L' + str(row)].value
    total = ws['N' + str(row)].value

    if client != 'Korporatīvais' or quantity is None or total is None:
        continue

    try:
        quantity = float(quantity)
        if 40 <= quantity <= 50:
            total_sum += float(total)
    except ValueError:
        continue 
print(int(total_sum))
