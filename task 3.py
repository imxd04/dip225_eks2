from openpyxl import load_workbook

wb = load_workbook('sagatave_eksamenam.xlsx')
ws = wb['Lapa_0']
max_row = ws.max_row

count = 0

for row in range(2, max_row + 1):
    address = ws['D' + str(row)].value
    city = ws['E' + str(row)].value

    if isinstance(address, str) and 'Adulienas iela' in address and city in ['Valmiera', 'Saulkrasti']:
        count += 1

print(count)
